from bitstring import BitArray
from random import random,randint,shuffle
from functools import reduce
import openpyxl
from openpyxl import Workbook
global pos
pos = 0
global temp
temp= []
nombreArchivo = "DatosGenerados.xlsx"


def pairwise(iterable):
    "s -> (s0, s1), (s2, s3), (s4, s5), ..."
    a = iter(iterable)
    return zip(a, a)

def generarIndividuo():
    Valor1=0
    Valor2=0
    Valor3=0
    Valor4=0

    Valor1 = randint(0, 10);
    Valor4 = 10 - Valor1
    Valor2 = randint(0,Valor4);
    Valor4 = Valor4 - Valor2
    Valor3 = randint(0,Valor4); #3
    Valor4= 10-(Valor1+Valor2+Valor3)
    individuo = [Valor1,Valor2,Valor3,Valor4]
    shuffle(individuo)
    individuoBinario = BitArray()
    for x in individuo:
        bitarrTemp = BitArray(4)
        bitarrTemp.uint = x
        # print("bitarrTemp:")
        # print(bin(x))
        # print(bitarrTemp.bin)
        individuoBinario = individuoBinario + bitarrTemp
    return individuoBinario

ganancias = [
    [0,0.28,0.45,0.65,0.78,0.90,1.02,1.13,1.23,1.32,1.38,0,0,0,0,0],
    [0,0.25,0.41,0.55,0.65,0.75,0.80,0.85,0.88,0.90,0.90,0,0,0,0,0],
    [0,0.15,0.25,0.40,0.50,0.62,0.73,0.82,0.90,0.96,1.00,0,0,0,0,0],
    [0,0.20,0.33,0.42,0.48,0.53,0.56,0.58,0.60,0.60,0.60,0,0,0,0,0]
]

class Genotipo:
    def __init__(self,fDecodificar):
        self.genes = []
        self.decodificar = fDecodificar
        self.longitud = 0
    def agregarGene(self,longitud,fEvaluar):
        inicio = self.longitud
        self.longitud = self.longitud+longitud
        self.genes.append(Gene(inicio,self.longitud,fEvaluar,self.decodificar))
    def aptitud(self,individuo):
        ganancias = self.sumarGanancias(individuo)
        sumaValoresGenes = self.sumarValoresGenes(individuo)
        v = 10-sumaValoresGenes
        return ganancias/(500*v+1)
    def sumarValoresGenes(self,individuo):
        return reduce(lambda sum,gene: sum+gene.getValor(individuo),self.genes,0)
    def sumarGanancias(self,individuo):
        return reduce(lambda sum,gene: sum+gene.evaluar(individuo),self.genes,0)
class Gene:
    def __init__(self,inicio,fin,fEvaluar,fDecodificar):
        self.inicio = inicio
        self.fin = fin
        self.evaluar = lambda cadena:fEvaluar(self.getValor(cadena))
        self.decodificar = fDecodificar
    def getValor(self,cadena):
        return self.decodificar(cadena[self.inicio:self.fin])

class TipoDato:
    @classmethod
    def entero(self,bitarr):
        return bitarr.uint

class Poblacion:
    def __init__(self,size,genotipo):
        self.genotipo = genotipo
        self.size = size
        self.individuos = [generarIndividuo() for i in range(0,size)]
        self.funcion = lambda poblacion:reduce(self.funcionComparativa,self.individuos,self.individuos[0])
    def funcionComparativa(self,ind1,ind2):
        if (self.genotipo.aptitud(ind1)>self.genotipo.aptitud(ind2)):
            return ind1
        else:
            return ind2
    def mejorIndividuo(self):
        return self.funcion(self.individuos)
    @classmethod
    def generarIndividuoAleatorio(self,longitud):
        return BitArray([random()<.5 for x in range(0,longitud)])
class Torneo:
    def operar(self,poblacion):
        mejor = poblacion.mejorIndividuo()
        print("Mejor individuo antes de torneo:"+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        #Se guarda en el archivo Excel
        global temp
        temp+=[pos,mejor.bin,poblacion.genotipo.aptitud(mejor)]
        ganadores1 = self.competencia(poblacion)
        poblacion.individuos = ganadores1 + self.competencia(poblacion)
        mejor = poblacion.mejorIndividuo()
        print("Mejor individuo despues de torneo:"+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        #Se guarda en el archivo Excel
        temp+=[mejor.bin,poblacion.genotipo.aptitud(mejor)]
    def competencia(self,poblacion):
        shuffle(poblacion.individuos)
        ganadores = []
        for ind1,ind2 in pairwise(poblacion.individuos):
            # print("peleando: "+str(ind1)+","+str(ind2))
            if (poblacion.genotipo.aptitud(ind1)>poblacion.genotipo.aptitud(ind2)):
                # print("gano:"+str(ind1))
                ganadores.append(ind1)
            else:
                # print("gano:"+str(ind2))
                ganadores.append(ind2)
        return ganadores

class Mutacion:
    def __init__(self,porcentaje,noPuntos):
        self.porcentaje = porcentaje
        self.noPuntos = noPuntos
    def operar(self,poblacion):
        mejor = poblacion.mejorIndividuo()
        print("Mejor individuo antes de mutacion:"+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        global temp
        temp+=[mejor.bin,poblacion.genotipo.aptitud(mejor)]
        for ind1 in poblacion.individuos:
            if (random()<=self.porcentaje):
                puntos = self.puntos(poblacion.genotipo.longitud)
                self.mutar(ind1,puntos)
        mejor = poblacion.mejorIndividuo()
        print("Mejor individuo despues de mutacion:"+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        temp+=[mejor.bin,poblacion.genotipo.aptitud(mejor)]
        
    def mutar(self,original,puntos):
        for i in puntos:
            original[i] = not original[i]
    def puntos(self,longitud):
        puntos = set()
        while len(puntos)<self.noPuntos:
            puntos.add(randint(0,longitud-1))
        return puntos

class Cruza:
    def __init__(self,porcentaje):
        self.porcentaje = porcentaje
    def operar(self,poblacion):
        mejor = poblacion.mejorIndividuo()
        print("Mejor individuo antes de cruza:"+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        global temp
        temp+=[mejor.bin,poblacion.genotipo.aptitud(mejor)]
        hijos = []
        for ind1,ind2 in pairwise(poblacion.individuos):
            if (random()<=self.porcentaje):
                puntos = self.puntos(poblacion.genotipo.longitud)
                hijos.append(self.cruzar(ind1,ind2,puntos))
                hijos.append(self.cruzar(ind2,ind1,puntos))
            else:
                hijos.append(ind1)
                hijos.append(ind2)
        poblacion.individuos = hijos
        mejor = poblacion.mejorIndividuo()
        print("Mejor individuo despues de cruza:"+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        temp+=[mejor.bin,poblacion.genotipo.aptitud(mejor)]
    def cruzar(self,padre1,padre2,puntos):
        # print("padre1: "+str(padre1)+", padre2: "+str(padre2))
        hijo = padre1[:puntos[0]]+padre2[puntos[0]:puntos[1]]+padre1[puntos[1]:]
        # print("hijo: "+str(hijo))
        return hijo
    def puntos(self,longitudGenotipo):
        # longitud de genotipo debe ser mayor o igual a 3
        p1 = randint(1,longitudGenotipo-2)
        p2 = randint(p1+1,longitudGenotipo-1)
        return (p1,p2)


class AG:
    def __init__(self):
        self.operadores = []
    def agregarOperadorGenetico(self,operador):
        self.operadores.append(operador)
    def ejecutar(self,poblacion,iteraciones):
        for i in range(0,iteraciones):
            mejor = poblacion.mejorIndividuo().copy()
            self.iteracion(poblacion)
            poblacion.individuos[0] = mejor
            self.imprimirResIteracion(poblacion,i)
            global pos
            pos += 1
            global temp
            temp=[]
    def iteracion(self,poblacion):
        for operador in self.operadores:
            operador.operar(poblacion)

    @classmethod
    def imprimirResIteracion(self,poblacion,n):
        mejor = poblacion.mejorIndividuo()
        print("Iteracion "+str(n)+": "+str(mejor.bin)+", apt: "+str(poblacion.genotipo.aptitud(mejor)))
        global temp
        temp+=[mejor.bin,poblacion.genotipo.aptitud(mejor)]
        hoja.append(temp)
ag = AG()
wb= openpyxl.load_workbook(nombreArchivo)
hoja= wb["AG"]
ag.agregarOperadorGenetico(Torneo())
ag.agregarOperadorGenetico(Cruza(0.8))
ag.agregarOperadorGenetico(Mutacion(0.01,2))

genotipo = Genotipo(TipoDato.entero)
genotipo.agregarGene(4,lambda x: ganancias[0][x])
genotipo.agregarGene(4,lambda x: ganancias[1][x])
genotipo.agregarGene(4,lambda x: ganancias[2][x])
genotipo.agregarGene(4,lambda x: ganancias[3][x])

# print(genotipo.aptitud(BitArray('0b1000001000000000')))

poblacion = Poblacion(50,genotipo)
ag.ejecutar(poblacion,20)
calcularmerge=hoja.max_row+1
mergeconcat1="A"+str(calcularmerge)
mergeconcat2="M"+str(calcularmerge)
mergefinal=mergeconcat1+":"+mergeconcat2
hoja.append(("Aqui termina la ejecución",""," "," "))
hoja.merge_cells(mergefinal)
wb.save(nombreArchivo)
# Torneo().operar(poblacion)
# print(generarInhoja.append(("","a","b","c"))dividuo().bin)